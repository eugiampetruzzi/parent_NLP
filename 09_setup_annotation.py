"""
09_setup_annotation.py
Creates the human_annotation/ folder with:
  1. annotation_guide.xlsx  — instructions, rating rubric, worked examples
  2. texts_to_rate.xlsx     — sampled texts (blinded to numeric score), one row per text,
                              empty columns for rater to fill in
  3. texts_to_rate.csv      — same content, Google Sheets-friendly

Sampling strategy
-----------------
- All 9 ASR text items included
- Up to 8 texts per item, stratified across numeric score levels (0 / 1 / 2)
  so annotators see the full range of endorsement severity
- Numeric score is HIDDEN from annotators (separate column, not shown) so
  their rating is independent — that's the whole point of the validation
- A small "calibration set" of 5 anchor examples (with correct ratings provided)
  is embedded at the top to orient annotators before they begin
"""

import random
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

ROOT    = Path(__file__).parent
DATA_IN = ROOT / "data" / "merged_clean.csv"
ANN_DIR = ROOT / "human_annotation"
ANN_DIR.mkdir(exist_ok=True)

MISSING = {"", "nan", "none", "0.0", "1.0", "888", "888.0", "999", "999.0"}
clean = lambda v: "" if str(v).strip().lower() in MISSING else str(v).strip()

random.seed(42)
np.random.seed(42)

ITEMS = {
    "asr_9_text.T1":   ("asr_9.T1",   "I can't get my mind off certain thoughts (describe):"),
    "asr_29_text.T1":  ("asr_29.T1",  "I am afraid of certain animals, situations, or places (describe):"),
    "asr_40_text.T1":  ("asr_40.T1",  "I hear sounds or voices that other people think aren't there (describe):"),
    "asr_46_text.T1":  ("asr_46.T1",  "Parts of my body twitch or make nervous movements (describe):"),
    "asr_58_text.T1":  ("asr_58.T1",  "I pick my skin or other parts of my body (describe):"),
    "asr_84_text.T1":  ("asr_84.T1",  "I do things that other people think are strange (describe):"),
    "asr_85_text.T1":  ("asr_85.T1",  "I have thoughts that other people would think are strange (describe):"),
    "asr_92_text.T1":  ("asr_92.T1",  "I do things that may cause me trouble with the law (describe):"),
    "asr_100_text.T1": ("asr_100.T1", "I have trouble sleeping (describe):"),
}

ITEM_LABELS = {
    "asr_9":   "Can't get mind off thoughts",
    "asr_29":  "Fears / phobias",
    "asr_40":  "Hears sounds / voices",
    "asr_46":  "Body twitches / tics",
    "asr_58":  "Skin picking",
    "asr_84":  "Does strange things",
    "asr_85":  "Strange thoughts",
    "asr_92":  "Trouble with law",
    "asr_100": "Trouble sleeping",
}

# Hard-coded anchor examples shown to annotators (with correct ratings given)
# These orient raters before they start blind coding
ANCHORS = [
    # (item_short, question_short, text, correct_severity, rationale)
    ("asr_29", "Fears / phobias",
     "spiders",
     0,
     "Single word, common mild fear, no indication of distress or life disruption."),
    ("asr_29", "Fears / phobias",
     "I have social anxiety, so social situations are stressful and scary for me.",
     2,
     "Explicitly names anxiety disorder, describes ongoing distress; suggests significant impairment."),
    ("asr_100", "Trouble sleeping",
     "occasionally",
     0,
     "Vague, minimal — only one word; no indication of severity or impact."),
    ("asr_100", "Trouble sleeping",
     "I fall asleep easily but I wake up in the middle of the night and can't get back to sleep.",
     2,
     "Specific pattern (sleep maintenance), nightly occurrence described, implies chronic impairment."),
    ("asr_9", "Can't get mind off thoughts",
     "worrying about money sometimes",
     1,
     "Some endorsement (not vague), but qualified ('sometimes') — moderate, not severe."),
]

# ── Load + sample ──────────────────────────────────────────────────────────
df = pd.read_csv(DATA_IN)

rows = []
text_id = 1

for text_col, (score_col, question) in ITEMS.items():
    item_short = text_col.replace("_text.T1", "")
    label = ITEM_LABELS.get(item_short, item_short)

    sub = df[[text_col, score_col]].copy()
    sub["text"]  = sub[text_col].apply(clean)
    sub["score"] = pd.to_numeric(sub[score_col], errors="coerce").replace({888: np.nan, 999: np.nan})
    sub = sub[sub["text"].str.len() > 2].dropna(subset=["score"])
    sub["score"] = sub["score"].astype(int)

    # Stratified sample: up to 3 per score level (0, 1, 2)
    sampled = []
    for level in [0, 1, 2]:
        pool = sub[sub["score"] == level]["text"].tolist()
        random.shuffle(pool)
        sampled.extend(pool[:3])

    # If fewer than 6 total, just take all
    if len(sampled) < 4:
        sampled = sub["text"].tolist()

    for t in sampled[:8]:
        rows.append({
            "text_id":     text_id,
            "item":        item_short,
            "item_label":  label,
            "question":    question,
            "response":    t,
            "_true_score": sub[sub["text"] == t]["score"].iloc[0]
                           if len(sub[sub["text"] == t]) > 0 else -1,
        })
        text_id += 1

main_df = pd.DataFrame(rows)
print(f"Total texts to annotate: {len(main_df)}")
print(main_df.groupby("item_label")["text_id"].count().to_string())

# ══════════════════════════════════════════════════════════════════════════
# Style helpers
# ══════════════════════════════════════════════════════════════════════════
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ORANGE     = "F4B942"
GREEN      = "70AD47"
RED        = "FF0000"
GRAY       = "F2F2F2"
WHITE      = "FFFFFF"

def hdr(cell, text, bg=DARK_BLUE, fg=WHITE, sz=12, bold=True, wrap=False, center=True):
    cell.value = text
    cell.font  = Font(bold=bold, size=sz, color=fg, name="Calibri")
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap)

def body(cell, text, bg=WHITE, sz=10, bold=False, wrap=True, left=True):
    cell.value = text
    cell.font  = Font(bold=bold, size=sz, name="Calibri")
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="top", wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, row, col):
    ws.cell(row=row, column=col).border = thin_border()

# ══════════════════════════════════════════════════════════════════════════
# Workbook 1 — ANNOTATION GUIDE
# ══════════════════════════════════════════════════════════════════════════
wb_guide = Workbook()
ws = wb_guide.active
ws.title = "Annotation Guide"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 80

def guide_title(ws, row, text):
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1)
    hdr(c, text, bg=DARK_BLUE, sz=14)
    ws.row_dimensions[row].height = 28

def guide_section(ws, row, text):
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1)
    hdr(c, text, bg=MID_BLUE, sz=11, center=False)
    ws.row_dimensions[row].height = 22

def guide_row(ws, row, label, value, bg=WHITE):
    la = ws.cell(row=row, column=1)
    va = ws.cell(row=row, column=2)
    body(la, label, bg=LIGHT_BLUE if bg == WHITE else bg, bold=True, sz=10)
    body(va, value, bg=bg, sz=10)
    ws.row_dimensions[row].height = max(15, min(80, 15 + value.count("\n") * 14))
    for c in [1, 2]:
        ws.cell(row=row, column=c).border = thin_border()

r = 1

guide_title(ws, r, "Human Annotation Guide — Parent ASR Free-Text Responses"); r += 2

guide_section(ws, r, "BACKGROUND"); r += 1
guide_row(ws, r, "What is this?",
    "Parents completed the Adult Self-Report (ASR), a standardized questionnaire "
    "about their own psychological functioning. For certain items, parents checked a "
    "0/1/2 rating AND wrote a free-text description of their experience.\n\n"
    "Your job is to read those descriptions and assign a severity rating — independently "
    "of the number the parent actually checked."); r += 1
guide_row(ws, r, "Why does this matter?",
    "We want to know whether the written text captures severity differently than the "
    "checkbox. If trained raters reading only the text agree with the checkbox, that "
    "validates the numeric scale. If they disagree, the text may be capturing something "
    "the checkbox misses — which is scientifically interesting."); r += 1
guide_row(ws, r, "Sample",
    "~65 text responses drawn from 9 ASR items. Each response is one parent's "
    "description. Responses were sampled to cover the full range of endorsement levels."); r += 2

guide_section(ws, r, "THE RATING SCALE"); r += 1
guide_row(ws, r, "0 — Not / barely a problem",
    "The description is vague, minimal, or describes something that causes little "
    "to no distress and does not interfere with daily functioning.\n\n"
    "Examples: 'spiders', 'occasionally', 'not really', 'a little'", bg=GRAY); r += 1
guide_row(ws, r, "1 — Somewhat a problem",
    "The description suggests a real pattern or moderate distress, but not overwhelming. "
    "The person is aware of it and it affects them, but they can generally function.\n\n"
    "Examples: 'I worry about money sometimes', 'I have trouble falling asleep a few "
    "nights a week', 'dogs scare me but I manage'"); r += 1
guide_row(ws, r, "2 — Significant problem",
    "The description suggests chronic, severe, or impairing symptoms. Explicit distress, "
    "avoidance that disrupts daily life, or clear clinical-level presentation.\n\n"
    "Examples: 'I have social anxiety and can't go to parties', 'I wake up every night "
    "and can't fall back asleep for hours', 'I obsess about this for hours daily'", bg=GRAY); r += 2

guide_section(ws, r, "ADDITIONAL RATINGS"); r += 1
guide_row(ws, r, "Specificity (1–3)",
    "1 = Very vague or one word (e.g., 'spiders', 'sometimes')\n"
    "2 = Moderately specific — some detail but not vivid\n"
    "3 = Specific and detailed — describes frequency, triggers, impact, or history"); r += 1
guide_row(ws, r, "Distress (0/1)",
    "0 = No language suggesting personal distress or suffering\n"
    "1 = Text contains distress language (e.g., 'terrified', 'can't stop', 'worried', "
    "'affects my life', 'I struggle')"); r += 1
guide_row(ws, r, "Confidence (1–3)",
    "1 = Uncertain — response is too short/ambiguous to rate reliably\n"
    "2 = Fairly confident\n"
    "3 = Very confident"); r += 2

guide_section(ws, r, "CALIBRATION EXAMPLES (correct ratings provided)"); r += 1
guide_row(ws, r, "Column headers", "Item  |  Response  |  Correct severity  |  Rationale"); r += 1

for item_s, item_l, text, sev, rationale in ANCHORS:
    guide_row(ws, r, f"{item_l}",
              f'RESPONSE: "{text}"\n\nCORRECT SEVERITY: {sev}\nRATIONALE: {rationale}',
              bg=LIGHT_BLUE if sev == 2 else (GRAY if sev == 0 else WHITE)); r += 1
r += 1

guide_section(ws, r, "TIPS & COMMON MISTAKES"); r += 1
guide_row(ws, r, "Rate the text, not the item",
    "A parent who wrote 'spiders' for the fears item and a parent who wrote 'spiders' "
    "for a different item should get the same severity rating — you're rating what "
    "the description communicates, not your prior belief about the item."); r += 1
guide_row(ws, r, "Short ≠ always 0",
    "'Police' (for fears) is short but specific and concerning — rate 1 or 2. "
    "'Occasionally' (for sleep) is short and vague — rate 0."); r += 1
guide_row(ws, r, "Don't overthink",
    "If a response reads as clinical-level to you on first read, rate 2. "
    "Trust your first impression and use the Confidence column to flag uncertainty."); r += 1
guide_row(ws, r, "Missing / unclear",
    "If a response is completely uninterpretable or unintelligible, rate severity = -1 "
    "and confidence = 1 and note in the comments column."); r += 1

wb_guide.save(ANN_DIR / "annotation_guide.xlsx")
print(f"Saved: {ANN_DIR}/annotation_guide.xlsx")

# ══════════════════════════════════════════════════════════════════════════
# Workbook 2 — TEXTS TO RATE
# ══════════════════════════════════════════════════════════════════════════
wb_rate = Workbook()

# ── Sheet 1: Instructions (brief) ─────────────────────────────────────────
ws_instr = wb_rate.active
ws_instr.title = "Instructions"
ws_instr.sheet_view.showGridLines = False
ws_instr.column_dimensions["A"].width = 20
ws_instr.column_dimensions["B"].width = 90

r = 1
guide_title(ws_instr, r, "Annotation Task — Quick Reference"); r += 2
guide_section(ws_instr, r, "YOUR TASK"); r += 1
guide_row(ws_instr, r, "Step 1",
    "Read the 'annotation_guide.xlsx' file first — especially the CALIBRATION "
    "EXAMPLES section. Do not start rating until you've read those."); r += 1
guide_row(ws_instr, r, "Step 2",
    "Go to the 'Calibration' sheet and rate the 5 calibration texts yourself "
    "(before looking at the correct answers). This checks your understanding."); r += 1
guide_row(ws_instr, r, "Step 3",
    "Go to the 'Rate These Texts' sheet. For each row, fill in:\n"
    "  • severity_rating   (0, 1, or 2)\n"
    "  • specificity       (1, 2, or 3)\n"
    "  • distress          (0 or 1)\n"
    "  • confidence        (1, 2, or 3)\n"
    "  • notes             (optional — anything that made the rating hard)"); r += 1
guide_row(ws_instr, r, "Scale summary",
    "severity:    0 = not/barely a problem    1 = somewhat    2 = significant problem\n"
    "specificity: 1 = vague/one-word          2 = moderate    3 = specific/detailed\n"
    "distress:    0 = no distress language    1 = distress mentioned\n"
    "confidence:  1 = uncertain               2 = fairly sure 3 = very confident"); r += 1
guide_row(ws_instr, r, "Do NOT look at",
    "The 'true_score' column is hidden. Do not unhide it until you're fully done. "
    "Your rating should be based only on the text."); r += 2
guide_section(ws_instr, r, "CONTACT"); r += 1
guide_row(ws_instr, r, "Questions?",
    "Email [PI name] or message on Slack. "
    "Please complete all rows in one sitting if possible (~30–45 min total)."); r += 1

# ── Sheet 2: Calibration (5 examples, rater fills in, then checks) ────────
ws_cal = wb_rate.create_sheet("Calibration")
ws_cal.sheet_view.showGridLines = False
ws_cal.freeze_panes = "A3"

CAL_COLS = [
    ("A", 6,  "text_id"),
    ("B", 22, "item"),
    ("C", 55, "response"),
    ("D", 14, "your_severity\n(0 / 1 / 2)"),
    ("E", 14, "your_specificity\n(1 / 2 / 3)"),
    ("F", 14, "your_distress\n(0 / 1)"),
    ("G", 14, "your_confidence\n(1 / 2 / 3)"),
    ("H", 30, "correct_severity\n(see guide)"),
    ("I", 50, "rationale"),
]
for col_letter, width, header in CAL_COLS:
    ws_cal.column_dimensions[col_letter].width = width

ws_cal.merge_cells("A1:I1")
hdr(ws_cal["A1"], "CALIBRATION — Rate these 5 texts, then check the 'correct_severity' column", bg=ORANGE, fg="000000", sz=12)
ws_cal.row_dimensions[1].height = 24

for j, (_, _, h) in enumerate(CAL_COLS, 1):
    c = ws_cal.cell(row=2, column=j)
    hdr(c, h, bg=MID_BLUE, sz=10, wrap=True)
    ws_cal.row_dimensions[2].height = 32

for i, (item_s, item_l, text, sev, rationale) in enumerate(ANCHORS, 3):
    data = [i - 2, item_l, text, "", "", "", "", sev, rationale]
    for j, val in enumerate(data, 1):
        c = ws_cal.cell(row=i, column=j)
        body(c, val if val != "" else "", bg=GRAY if i % 2 == 0 else WHITE, sz=10)
        c.border = thin_border()
    ws_cal.row_dimensions[i].height = 55

# ── Sheet 3: Main annotation sheet ────────────────────────────────────────
ws_rate = wb_rate.create_sheet("Rate These Texts")
ws_rate.sheet_view.showGridLines = False
ws_rate.freeze_panes = "A3"

RATE_COLS = [
    ("A", 7,   "text_id"),
    ("B", 22,  "item"),
    ("C", 40,  "question"),
    ("D", 70,  "parent_response"),
    ("E", 16,  "severity_rating\n(0 / 1 / 2)"),
    ("F", 16,  "specificity\n(1 / 2 / 3)"),
    ("G", 14,  "distress\n(0 / 1)"),
    ("H", 14,  "confidence\n(1 / 2 / 3)"),
    ("I", 40,  "notes\n(optional)"),
    ("J", 12,  "_true_score\n(DO NOT LOOK)"),  # hidden
]
for col_letter, width, header in RATE_COLS:
    ws_rate.column_dimensions[col_letter].width = width

# Title row
ws_rate.merge_cells("A1:J1")
hdr(ws_rate["A1"],
    "Rate These Texts — fill in columns E–I for every row  (do not look at column J until done)",
    bg=DARK_BLUE, sz=12)
ws_rate.row_dimensions[1].height = 24

# Header row
for j, (_, _, h) in enumerate(RATE_COLS, 1):
    c = ws_rate.cell(row=2, column=j)
    hdr(c, h, bg=MID_BLUE, sz=10, wrap=True)
ws_rate.row_dimensions[2].height = 36

# Group rows by item, with a shaded separator between items
prev_item = None
data_row = 3
for _, row in main_df.iterrows():
    item = row["item_label"]
    is_new_group = item != prev_item
    bg = LIGHT_BLUE if is_new_group else (GRAY if data_row % 2 == 0 else WHITE)
    prev_item = item

    vals = [
        row["text_id"],
        row["item_label"],
        row["question"],
        row["response"],
        "",   # severity_rating
        "",   # specificity
        "",   # distress
        "",   # confidence
        "",   # notes
        row["_true_score"],  # hidden
    ]
    for j, val in enumerate(vals, 1):
        c = ws_rate.cell(row=data_row, column=j)
        body(c, val, bg=bg, sz=10)
        c.border = thin_border()

    # Make the rater-input cells stand out
    for j in [5, 6, 7, 8]:
        c = ws_rate.cell(row=data_row, column=j)
        c.fill = PatternFill("solid", fgColor="FFFDE7")  # pale yellow
        c.font = Font(bold=True, size=11, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws_rate.row_dimensions[data_row].height = max(40, min(100, 40 + len(row["response"]) // 45 * 14))
    data_row += 1

# Hide the true_score column
ws_rate.column_dimensions["J"].hidden = True

# ── Summary sheet ──────────────────────────────────────────────────────────
ws_sum = wb_rate.create_sheet("Summary Stats")
ws_sum.sheet_view.showGridLines = False
ws_sum.column_dimensions["A"].width = 28
ws_sum.column_dimensions["B"].width = 14

ws_sum.merge_cells("A1:B1")
hdr(ws_sum["A1"], "Item breakdown — texts included in annotation set", bg=DARK_BLUE, sz=12)
ws_sum.row_dimensions[1].height = 24

hdr(ws_sum.cell(2, 1), "Item", bg=MID_BLUE)
hdr(ws_sum.cell(2, 2), "N texts", bg=MID_BLUE)
ws_sum.row_dimensions[2].height = 20

counts = main_df.groupby("item_label")["text_id"].count()
for i, (lbl, n) in enumerate(counts.items(), 3):
    body(ws_sum.cell(i, 1), lbl, bg=GRAY if i % 2 == 0 else WHITE)
    body(ws_sum.cell(i, 2), int(n), bg=GRAY if i % 2 == 0 else WHITE, left=False)
    for col in [1, 2]:
        ws_sum.cell(i, col).border = thin_border()

total_row = 3 + len(counts)
hdr(ws_sum.cell(total_row, 1), "TOTAL", bg=DARK_BLUE)
hdr(ws_sum.cell(total_row, 2), len(main_df), bg=DARK_BLUE)

wb_rate.save(ANN_DIR / "texts_to_rate.xlsx")
print(f"Saved: {ANN_DIR}/texts_to_rate.xlsx")

# ── CSV backup ────────────────────────────────────────────────────────────
export = main_df[["text_id", "item_label", "question", "response"]].copy()
export["severity_rating"] = ""
export["specificity"]     = ""
export["distress"]        = ""
export["confidence"]      = ""
export["notes"]           = ""
export.to_csv(ANN_DIR / "texts_to_rate.csv", index=False)
print(f"Saved: {ANN_DIR}/texts_to_rate.csv")

# ── Print item distribution ────────────────────────────────────────────────
print("\nItem distribution:")
print(main_df.groupby(["item_label", "_true_score"])["text_id"].count()
      .unstack(fill_value=0).to_string())
print(f"\nTotal texts: {len(main_df)}")
print(f"\nAll files in: {ANN_DIR}/")
